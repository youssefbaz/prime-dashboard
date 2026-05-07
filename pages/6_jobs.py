import streamlit as st
import datetime
import json
import re
import io
import html as _html
from utils import load_data, save_data, get_week_info

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

data      = load_data()
wi        = get_week_info()
today_str = wi["today_str"]
today     = wi["today"]

# ── Claude API key ─────────────────────────────────────────
try:
    CLAUDE_KEY = st.secrets["ANTHROPIC_API_KEY"]
except Exception:
    CLAUDE_KEY = ""

# ── Excel columns spec ────────────────────────────────────
EXCEL_COLS = [
    ("Date Applied",   18),
    ("Company",        22),
    ("Role / Title",   28),
    ("Location",       20),
    ("Contract Type",  16),
    ("Salary",         16),
    ("Source",         16),
    ("Status",         14),
    ("URL",            40),
    ("Notes",          35),
]

STATUS_OPTIONS  = ["Applied", "Interview", "Rejected", "Offer", "Ghosted"]
CONTRACT_TYPES  = ["CDI", "CDD", "Internship", "Freelance", "Part-time", "Other"]
SOURCE_OPTIONS  = ["LinkedIn", "Indeed", "Glassdoor", "Company Site", "Referral", "Other"]

STATUS_COLORS = {
    "Applied":   "#818cf8",
    "Interview": "#34d399",
    "Rejected":  "#ef4444",
    "Offer":     "#f472b6",
    "Ghosted":   "#64748b",
}

# ── Helpers ────────────────────────────────────────────────
def fetch_url_text(url: str) -> str:
    """Fetch a URL and return readable text (stripped HTML)."""
    if not HAS_REQUESTS:
        return ""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; PrimeDashboard/1.0)"}
        r = requests.get(url, headers=headers, timeout=15)
        r.raise_for_status()
        if HAS_BS4:
            soup = BeautifulSoup(r.text, "html.parser")
            for tag in soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            return soup.get_text(separator="\n", strip=True)[:6000]
        return r.text[:6000]
    except Exception as e:
        return f"ERROR: {e}"


def extract_job_info(text: str) -> dict:
    """Call Claude API to extract structured job info from raw text."""
    if not CLAUDE_KEY or not HAS_REQUESTS:
        return {}
    prompt = f"""Extract job application details from the text below and return ONLY valid JSON with these exact keys:
company, title, location, contract_type, salary, source

Rules:
- contract_type must be one of: CDI, CDD, Internship, Freelance, Part-time, Other
- source must be one of: LinkedIn, Indeed, Glassdoor, Company Site, Referral, Other
- salary: include currency and period if visible (e.g. "45 000 € / year"), else ""
- location: city + country if available, else ""
- If a field cannot be determined, use ""
- Return ONLY the JSON object, no extra text

TEXT:
{text[:5000]}"""

    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "x-api-key": CLAUDE_KEY,
                "anthropic-version": "2023-06-01",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 512,
                "system": "You extract structured job data. Return only valid JSON, no markdown, no extra text.",
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        if r.status_code == 200:
            raw = r.json()["content"][0]["text"].strip()
            raw = re.sub(r"```json|```", "", raw).strip()
            s = raw.find("{"); e = raw.rfind("}") + 1
            return json.loads(raw[s:e])
    except Exception:
        pass
    return {}


def build_excel(jobs: list) -> bytes:
    """Build a styled Excel workbook from the jobs list and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1E1B4B")
    header_font = Font(bold=True, color="C7D2FE", size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
    )

    status_fill = {
        "Applied":   "1E1B4B",
        "Interview": "064E3B",
        "Rejected":  "450A0A",
        "Offer":     "500724",
        "Ghosted":   "1E293B",
    }
    status_font_color = {
        "Applied":   "818CF8",
        "Interview": "34D399",
        "Rejected":  "F87171",
        "Offer":     "F472B6",
        "Ghosted":   "64748B",
    }

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(EXCEL_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Write rows
    row_fill_light = PatternFill("solid", fgColor="0F172A")
    row_fill_dark  = PatternFill("solid", fgColor="0A0A0F")

    for row_idx, job in enumerate(jobs, start=2):
        fill = row_fill_light if row_idx % 2 == 0 else row_fill_dark
        default_font = Font(color="CBD5E1", size=10)

        values = [
            job.get("date", ""),
            job.get("company", ""),
            job.get("role", ""),
            job.get("location", ""),
            job.get("contract_type", ""),
            job.get("salary", ""),
            job.get("source", ""),
            job.get("status", "Applied"),
            job.get("url", ""),
            job.get("note", ""),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 9, 10)))

            # Special status column styling
            col_name = EXCEL_COLS[col_idx - 1][0]
            if col_name == "Status":
                s = str(val)
                cell.font = Font(
                    bold=True,
                    color=status_font_color.get(s, "94A3B8"),
                    size=10
                )
                cell.fill = PatternFill("solid", fgColor=status_fill.get(s, "1E293B"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "URL" and val:
                cell.font = Font(color="6366F1", size=10, underline="single")
                cell.hyperlink = str(val)
            else:
                cell.font = default_font

        ws.row_dimensions[row_idx].height = 20

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLS))}{len(jobs) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════
# PAGE
# ══════════════════════════════════════════════════════════
st.markdown('<p class="page-title">Job Tracker</p>', unsafe_allow_html=True)
st.markdown('<p class="page-sub">Track every application — let AI do the data entry.</p>', unsafe_allow_html=True)

jobs = data.get("jobs", [])

# Stats bar ──────────────────────────────────────────────
if jobs:
    status_counts = {}
    for j in jobs:
        status_counts[j.get("status", "Applied")] = status_counts.get(j.get("status", "Applied"), 0) + 1

    this_week = sum(1 for j in jobs if j.get("date","") >= (today - datetime.timedelta(days=7)).isoformat())
    last_week = sum(1 for j in jobs if (today - datetime.timedelta(days=14)).isoformat()
                   <= j.get("date","") < (today - datetime.timedelta(days=7)).isoformat())
    wk_arrow  = "▲" if this_week >= last_week else "▼"
    wk_color  = "#34d399" if this_week >= last_week else "#ef4444"
    total     = len(jobs)
    ir        = round(status_counts.get("Interview", 0) / total * 100) if total else 0

    sc1, sc2, sc3, sc4, sc5 = st.columns(5)
    with sc1:
        st.markdown(f'<div class="s-card"><div class="s-label">Total</div><div class="s-val" style="color:#818cf8;">{total}</div></div>', unsafe_allow_html=True)
    with sc2:
        st.markdown(f'<div class="s-card"><div class="s-label">This week</div><div class="s-val" style="color:{wk_color};">{this_week} <span style="font-size:14px;">{wk_arrow}</span></div><div class="s-sub">last: {last_week}</div></div>', unsafe_allow_html=True)
    with sc3:
        st.markdown(f'<div class="s-card"><div class="s-label">Interviews</div><div class="s-val" style="color:#34d399;">{status_counts.get("Interview",0)}</div></div>', unsafe_allow_html=True)
    with sc4:
        st.markdown(f'<div class="s-card"><div class="s-label">Offers</div><div class="s-val" style="color:#f472b6;">{status_counts.get("Offer",0)}</div></div>', unsafe_allow_html=True)
    with sc5:
        st.markdown(f'<div class="s-card"><div class="s-label">Interview rate</div><div class="s-val" style="color:#fbbf24;">{ir}%</div></div>', unsafe_allow_html=True)

    st.markdown("")

# Tabs ───────────────────────────────────────────────────
tab_smart, tab_manual, tab_all = st.tabs(["✨ Smart Add (AI)", "✏️ Manual Add", "📋 All Applications"])

# ── TAB 1: Smart Add ──────────────────────────────────────
with tab_smart:
    if not CLAUDE_KEY:
        st.error("Add `ANTHROPIC_API_KEY` to `.streamlit/secrets.toml` to use Smart Add.")
    else:
        st.markdown('<div class="sec-title">Paste job offer or URL</div>', unsafe_allow_html=True)

        input_mode = st.radio(
            "Input type",
            ["Paste URL", "Paste job text / description"],
            horizontal=True,
            key="smart_mode",
        )

        if input_mode == "Paste URL":
            job_input = st.text_input(
                "Job posting URL",
                placeholder="https://www.linkedin.com/jobs/view/...",
                key="smart_url_input",
            )
        else:
            job_input = st.text_area(
                "Job offer text",
                placeholder="Paste the full job title, description, or ad here…",
                height=180,
                key="smart_text_input",
            )

        extract_btn = st.button("🤖 Extract with AI", use_container_width=True, type="primary", key="extract_btn")

        if extract_btn and job_input.strip():
            with st.spinner("Fetching and analysing…"):
                if input_mode == "Paste URL":
                    raw_text = fetch_url_text(job_input.strip())
                    if raw_text.startswith("ERROR:"):
                        st.warning(f"Could not fetch URL: {raw_text}. Tip: paste the job text directly instead.")
                        raw_text = ""
                    source_url = job_input.strip()
                else:
                    raw_text   = job_input.strip()
                    source_url = ""

                if raw_text:
                    extracted = extract_job_info(raw_text)
                    st.session_state["smart_extracted"] = extracted
                    st.session_state["smart_url"]       = source_url

        # Review & edit form
        if "smart_extracted" in st.session_state:
            ex  = st.session_state["smart_extracted"]
            url = st.session_state.get("smart_url", "")

            st.markdown("")
            st.markdown('<div class="sec-title">Review & edit</div>', unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:13px;color:#64748b;margin-bottom:16px;">'
                'Fix anything the AI got wrong, then save.</div>',
                unsafe_allow_html=True,
            )

            r1c1, r1c2 = st.columns(2)
            with r1c1:
                s_company = st.text_input("Company *", value=ex.get("company", ""), key="s_company")
            with r1c2:
                s_role = st.text_input("Role / Title *", value=ex.get("title", ""), key="s_role")

            r2c1, r2c2, r2c3 = st.columns(3)
            with r2c1:
                s_location = st.text_input("Location", value=ex.get("location", ""), key="s_location")
            with r2c2:
                ct_default = ex.get("contract_type", "CDI")
                ct_idx = CONTRACT_TYPES.index(ct_default) if ct_default in CONTRACT_TYPES else 0
                s_contract = st.selectbox("Contract type", CONTRACT_TYPES, index=ct_idx, key="s_contract")
            with r2c3:
                s_salary = st.text_input("Salary", value=ex.get("salary", ""), key="s_salary")

            r3c1, r3c2 = st.columns(2)
            with r3c1:
                src_default = ex.get("source", "Other")
                src_idx = SOURCE_OPTIONS.index(src_default) if src_default in SOURCE_OPTIONS else len(SOURCE_OPTIONS) - 1
                s_source = st.selectbox("Source", SOURCE_OPTIONS, index=src_idx, key="s_source")
            with r3c2:
                s_status = st.selectbox("Status", STATUS_OPTIONS, index=0, key="s_status")

            s_url  = st.text_input("URL", value=url, key="s_url")
            s_note = st.text_input("Notes (optional)", key="s_note")

            col_save, col_clear = st.columns([3, 1])
            with col_save:
                if st.button("💾 Save to tracker", use_container_width=True, type="primary", key="save_smart"):
                    if not s_company.strip() or not s_role.strip():
                        st.error("Company and Role are required.")
                    else:
                        new_job = {
                            "date":          today_str,
                            "company":       s_company.strip(),
                            "role":          s_role.strip(),
                            "location":      s_location.strip(),
                            "contract_type": s_contract,
                            "salary":        s_salary.strip(),
                            "source":        s_source,
                            "status":        s_status,
                            "url":           s_url.strip(),
                            "note":          s_note.strip(),
                        }
                        data.setdefault("jobs", []).append(new_job)
                        save_data(data)
                        del st.session_state["smart_extracted"]
                        if "smart_url" in st.session_state:
                            del st.session_state["smart_url"]
                        st.success(f"Saved — {s_company.strip()} · {s_role.strip()}")
                        st.rerun()
            with col_clear:
                if st.button("✕ Clear", use_container_width=True, key="clear_smart"):
                    del st.session_state["smart_extracted"]
                    if "smart_url" in st.session_state:
                        del st.session_state["smart_url"]
                    st.rerun()

# ── TAB 2: Manual Add ─────────────────────────────────────
with tab_manual:
    st.markdown('<div class="sec-title">Add application manually</div>', unsafe_allow_html=True)
    with st.form("job_form_manual", clear_on_submit=True):
        mc1, mc2 = st.columns(2)
        with mc1:
            m_company = st.text_input("Company *")
        with mc2:
            m_role = st.text_input("Role / Title *")

        mc3, mc4, mc5 = st.columns(3)
        with mc3:
            m_location = st.text_input("Location")
        with mc4:
            m_contract = st.selectbox("Contract type", CONTRACT_TYPES)
        with mc5:
            m_salary = st.text_input("Salary")

        mc6, mc7 = st.columns(2)
        with mc6:
            m_source = st.selectbox("Source", SOURCE_OPTIONS)
        with mc7:
            m_status = st.selectbox("Status", STATUS_OPTIONS)

        m_url  = st.text_input("URL (optional)")
        m_note = st.text_input("Notes (optional)")

        if st.form_submit_button("➕ Add application", use_container_width=True):
            m_company = m_company.strip()
            m_role    = m_role.strip()
            if m_company and m_role:
                data.setdefault("jobs", []).append({
                    "date":          today_str,
                    "company":       m_company,
                    "role":          m_role,
                    "location":      m_location.strip(),
                    "contract_type": m_contract,
                    "salary":        m_salary.strip(),
                    "source":        m_source,
                    "status":        m_status,
                    "url":           m_url.strip(),
                    "note":          m_note.strip(),
                })
                save_data(data)
                st.success(f"Added — {m_company} · {m_role}")
                st.rerun()
            elif not m_company:
                st.error("Company name is required.")
            else:
                st.error("Role / Title is required.")

# ── TAB 3: All Applications ───────────────────────────────
with tab_all:
    jobs = data.get("jobs", [])

    if not jobs:
        st.info("No applications yet. Use Smart Add or Manual Add to get started!")
    else:
        # Filter + Excel download row
        fc1, fc2 = st.columns([3, 1])
        with fc1:
            filter_status = st.selectbox(
                "Filter by status",
                ["All"] + STATUS_OPTIONS,
                key="job_filter",
            )
        with fc2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if HAS_OPENPYXL:
                excel_bytes = build_excel(jobs)
                filename    = f"job_tracker_{today_str}.xlsx"
                st.download_button(
                    label="⬇ Download Excel",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_excel",
                )
            else:
                st.warning("openpyxl not installed.")

        filtered = jobs if filter_status == "All" else [j for j in jobs if j.get("status") == filter_status]

        st.markdown(
            f'<div style="font-size:13px;color:#64748b;margin:8px 0 16px;">'
            f'Showing {len(filtered)} of {len(jobs)} application(s)</div>',
            unsafe_allow_html=True,
        )

        display_jobs = list(reversed(filtered[-50:]))
        for idx, j in enumerate(display_jobs):
            # find the true index in data["jobs"] for editing
            global_idx = data["jobs"].index(j) if j in data["jobs"] else -1

            sc     = STATUS_COLORS.get(j.get("status", "Applied"), "#64748b")
            url_lk = (
                f' <a href="{_html.escape(j["url"])}" target="_blank" '
                f'style="font-size:11px;color:#475569;text-decoration:none;">↗</a>'
                if j.get("url") else ""
            )
            cl_badge = (
                '<span style="font-size:10px;color:#818cf8;background:rgba(129,140,248,0.12);'
                'padding:2px 8px;border-radius:4px;margin-left:8px;">📝 cover letter</span>'
                if j.get("has_cover_letter") else ""
            )

            meta_parts = []
            if j.get("location"):      meta_parts.append(f"📍 {_html.escape(j['location'])}")
            if j.get("contract_type"): meta_parts.append(f"📄 {_html.escape(j['contract_type'])}")
            if j.get("salary"):        meta_parts.append(f"💰 {_html.escape(j['salary'])}")
            if j.get("source"):        meta_parts.append(f"🔗 {_html.escape(j['source'])}")
            meta_line = (
                f'<div style="font-size:11px;color:#475569;margin-top:3px;">'
                + " &nbsp;·&nbsp; ".join(meta_parts) + "</div>"
                if meta_parts else ""
            )
            note_line = (
                f'<div style="font-size:12px;color:#475569;margin-top:2px;">{_html.escape(j["note"])}</div>'
                if j.get("note") else ""
            )

            row_col, status_col = st.columns([5, 1])
            with row_col:
                st.markdown(f"""
<div class="job-row" style="margin-bottom:0;">
  <div style="flex:1;min-width:0;">
    <div>
      <span style="font-weight:600;color:#e2e8f0;font-size:14px;">{_html.escape(j.get('company',''))}</span>
      <span style="color:#64748b;margin-left:8px;font-size:13px;">{_html.escape(j.get('role',''))}</span>
      {cl_badge}{url_lk}
    </div>
    {meta_line}
    {note_line}
  </div>
  <div style="display:flex;align-items:center;gap:12px;flex-shrink:0;">
    <span style="font-size:12px;color:#475569;font-family:'JetBrains Mono',monospace;">{j.get('date','')}</span>
  </div>
</div>""", unsafe_allow_html=True)
            with status_col:
                current_status = j.get("status", "Applied")
                new_status = st.selectbox(
                    "Status",
                    STATUS_OPTIONS,
                    index=STATUS_OPTIONS.index(current_status) if current_status in STATUS_OPTIONS else 0,
                    key=f"status_sel_{global_idx}_{idx}",
                    label_visibility="collapsed",
                )
                if new_status != current_status and global_idx >= 0:
                    data["jobs"][global_idx]["status"] = new_status
                    save_data(data)
                    st.rerun()

st.markdown('<div class="prime-footer">Every application is a vote for the future you want. 💼</div>', unsafe_allow_html=True)
